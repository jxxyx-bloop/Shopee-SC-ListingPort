"""Write transformed product data into Shopee Mass Upload template."""

from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
import pandas as pd


def read_template_columns(template_path: str | Path) -> list[str]:
    """Read the column names from row 2 (0-indexed) of the Template sheet.

    Uses calamine engine since openpyxl can't read Shopee-generated files.
    """
    df = pd.read_excel(
        template_path, engine="calamine", sheet_name="Template", header=None
    )
    # Row 2 (0-indexed) contains the human-readable column names
    columns = [str(v) if pd.notna(v) else "" for v in df.iloc[2].tolist()]
    return columns


def _patch_and_load_template(template_path: str | Path) -> openpyxl.Workbook:
    """Load a Shopee template xlsx, working around the activePane bug.

    Shopee files have a non-standard activePane value that crashes openpyxl.
    We read the raw XML, fix the issue, and load the workbook.
    """
    import io
    import re
    import zipfile

    with open(template_path, "rb") as f:
        data = f.read()

    # Read the xlsx (which is a zip file), fix any problematic sheet XML
    buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data), "r") as zin:
        with zipfile.ZipFile(buf, "w") as zout:
            for item in zin.infolist():
                content = zin.read(item.filename)
                if item.filename.startswith("xl/worksheets/"):
                    # Remove or fix invalid activePane attributes
                    content_str = content.decode("utf-8")
                    # Remove activePane="" (empty value that breaks openpyxl)
                    content_str = re.sub(r'activePane="[^"]*"', "", content_str)
                    content = content_str.encode("utf-8")
                zout.writestr(item, content)

    buf.seek(0)
    return openpyxl.load_workbook(buf)


def write_upload_file(
    rows: list[dict],
    template_path: str | Path,
    output_path: str | Path,
) -> Path:
    """Write transformed product rows into a copy of the upload template.

    Args:
        rows: List of dicts from mapper.transform_products()
        template_path: Path to the blank upload template xlsx
        output_path: Path for the output xlsx file

    Returns:
        Path to the written output file.
    """
    output_path = Path(output_path)

    # Read column names from template (using calamine for reading)
    template_columns = read_template_columns(template_path)

    # Load template with openpyxl for writing (with activePane fix)
    wb = _patch_and_load_template(template_path)
    ws = wb["Template"]

    # Data starts at row 6 (1-indexed), i.e., after the 5 header rows (rows 1-5)
    data_start_row = 6

    # Clear any existing data rows (in case template has sample data)
    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - data_start_row + 1)

    # Write each product row
    for row_idx, row_data in enumerate(rows):
        excel_row = data_start_row + row_idx
        for col_idx, col_name in enumerate(template_columns):
            if not col_name:
                continue
            excel_col = col_idx + 1  # openpyxl is 1-indexed
            value = row_data.get(col_name)
            if value is not None:
                ws.cell(row=excel_row, column=excel_col, value=value)

    wb.save(output_path)
    return output_path
